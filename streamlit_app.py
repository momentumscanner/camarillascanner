
import streamlit as st
import pandas as pd
import io
import os
import re
from scanner import CamarillaScanner
from openpyxl.styles import Alignment, Font, PatternFill

# Page configuration
st.set_page_config(
    page_title="Camarilla Option Scanner",
    page_icon="📈",
    layout="wide"
)

# Custom CSS
st.markdown("""
    <style>
    .main {
        padding: 2rem;
    }
    .stButton>button {
        width: 100%;
        background-color: #27ae60;
        color: white;
        height: 3em;
        font-weight: bold;
    }
    .success-msg {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #d4edda;
        color: #155724;
        margin-bottom: 1rem;
    }
    </style>
""", unsafe_allow_html=True)

def generate_excel(df, today_filename, top_n=5):
    """
    Generates the Excel report in memory and returns the BytesIO object.
    Preserves the formatting logic from the desktop app.
    """
    output = io.BytesIO()
    
    # Reorder columns
    cols = list(df.columns)
    priority = ['Symbol', 'Expiry', 'ATM_Strike', 'Option_Type', 'Spot_Close']
    final_cols = priority + [c for c in cols if c not in priority]
    df_final = df[final_cols]
    
    # Filter for Sheet 2
    df_inside_full = df[df['Is_Inside_Camarilla'] == True].copy()
    
    # Split CE and PE
    cols_to_show = ['Symbol', 'Spot_Close', 'ATM_Strike']
    
    df_ce = df_inside_full[df_inside_full['Option_Type'] == 'CE'][cols_to_show].copy()
    df_pe = df_inside_full[df_inside_full['Option_Type'] == 'PE'][cols_to_show].copy()
    
    df_ce.reset_index(drop=True, inplace=True)
    df_pe.reset_index(drop=True, inplace=True)
    
    df_combined = pd.concat([df_ce, df_pe], axis=1)
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Main Data
        df_final.to_excel(writer, sheet_name='Main Data', index=False)
        
        # Sheet 2: Inside Camarilla
        sheet_name = 'Narrow Camarilla'
        df_combined.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Header Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center = Alignment(horizontal='center', vertical='center')
        
        # Merge and Write CE Header
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        cell_ce = worksheet.cell(row=1, column=1)
        cell_ce.value = "Narrow Camarilla CE"
        cell_ce.font = header_font
        cell_ce.fill = header_fill
        cell_ce.alignment = center
        
        # Merge and Write PE Header
        worksheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)
        cell_pe = worksheet.cell(row=1, column=4)
        cell_pe.value = "Narrow Camarilla PE"
        cell_pe.font = header_font
        cell_pe.fill = header_fill
        cell_pe.alignment = center
        
        # Formatting columns width
        for i, col in enumerate(df_combined.columns):
            col_idx = i + 1 
            series = df_combined.iloc[:, i]
            max_len = max((series.apply(str).map(len).max() if not series.empty else 0), len(str(col))) + 2
            column_letter = worksheet.cell(row=2, column=col_idx).column_letter
            worksheet.column_dimensions[column_letter].width = max_len

        # Sheet 3: Inside Camarilla (H4/L4 Logic)
        df_h4_l4_full = df[df['Is_Inside_H4_L4'] == True].copy()
        
        df_ce_h4 = df_h4_l4_full[df_h4_l4_full['Option_Type'] == 'CE'][cols_to_show].copy()
        df_pe_h4 = df_h4_l4_full[df_h4_l4_full['Option_Type'] == 'PE'][cols_to_show].copy()
        
        df_ce_h4.reset_index(drop=True, inplace=True)
        df_pe_h4.reset_index(drop=True, inplace=True)
        
        df_combined_h4 = pd.concat([df_ce_h4, df_pe_h4], axis=1)
        
        sheet_name_h4 = 'Inside Camarilla'
        df_combined_h4.to_excel(writer, sheet_name=sheet_name_h4, index=False, startrow=1)
        
        worksheet_h4 = writer.sheets[sheet_name_h4]
        
        # Headers for Sheet 3
        worksheet_h4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        cell_ce_h4 = worksheet_h4.cell(row=1, column=1)
        cell_ce_h4.value = "Inside Camarilla CE"
        cell_ce_h4.font = header_font
        cell_ce_h4.fill = header_fill
        cell_ce_h4.alignment = center
        
        worksheet_h4.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)
        cell_pe_h4 = worksheet_h4.cell(row=1, column=4)
        cell_pe_h4.value = "Inside Camarilla PE"
        cell_pe_h4.font = header_font
        cell_pe_h4.fill = header_fill
        cell_pe_h4.alignment = center
        
        # Formatting columns for Sheet 3
        for i, col in enumerate(df_combined_h4.columns):
            col_idx = i + 1
            series = df_combined_h4.iloc[:, i]
            max_len = max((series.apply(str).map(len).max() if not series.empty else 0), len(str(col))) + 2
            column_letter = worksheet_h4.cell(row=2, column=col_idx).column_letter
            worksheet_h4.column_dimensions[column_letter].width = max_len

        # Sheet 4: Higher Value Camarilla
        df_higher_full = df[df['Is_Higher_Value'] == True].copy()
        
        df_ce_higher = df_higher_full[df_higher_full['Option_Type'] == 'CE'][cols_to_show].copy()
        df_pe_higher = df_higher_full[df_higher_full['Option_Type'] == 'PE'][cols_to_show].copy()
        
        df_ce_higher.reset_index(drop=True, inplace=True)
        df_pe_higher.reset_index(drop=True, inplace=True)
        
        df_combined_higher = pd.concat([df_ce_higher, df_pe_higher], axis=1)
        
        sheet_name_higher = 'Higher Value Camarilla'
        df_combined_higher.to_excel(writer, sheet_name=sheet_name_higher, index=False, startrow=1)
        
        worksheet_higher = writer.sheets[sheet_name_higher]
        
        # Headers
        worksheet_higher.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        cell_ce_h = worksheet_higher.cell(row=1, column=1)
        cell_ce_h.value = "Higher Value Camarilla CE"
        cell_ce_h.font = header_font
        cell_ce_h.fill = header_fill
        cell_ce_h.alignment = center
        
        worksheet_higher.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)
        cell_pe_h = worksheet_higher.cell(row=1, column=4)
        cell_pe_h.value = "Higher Value Camarilla PE"
        cell_pe_h.font = header_font
        cell_pe_h.fill = header_fill
        cell_pe_h.alignment = center
        
        # Formatting
        for i, col in enumerate(df_combined_higher.columns):
            col_idx = i + 1
            series = df_combined_higher.iloc[:, i]
            max_len = max((series.apply(str).map(len).max() if not series.empty else 0), len(str(col))) + 2
            column_letter = worksheet_higher.cell(row=2, column=col_idx).column_letter
            worksheet_higher.column_dimensions[column_letter].width = max_len

        # Sheet 5: Lower Value Camarilla
        df_lower_full = df[df['Is_Lower_Value'] == True].copy()
        
        df_ce_lower = df_lower_full[df_lower_full['Option_Type'] == 'CE'][cols_to_show].copy()
        df_pe_lower = df_lower_full[df_lower_full['Option_Type'] == 'PE'][cols_to_show].copy()
        
        df_ce_lower.reset_index(drop=True, inplace=True)
        df_pe_lower.reset_index(drop=True, inplace=True)
        
        df_combined_lower = pd.concat([df_ce_lower, df_pe_lower], axis=1)
        
        sheet_name_lower = 'Lower Value Camarilla'
        df_combined_lower.to_excel(writer, sheet_name=sheet_name_lower, index=False, startrow=1)
        
        worksheet_lower = writer.sheets[sheet_name_lower]
        
        # Headers
        worksheet_lower.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        cell_ce_l = worksheet_lower.cell(row=1, column=1)
        cell_ce_l.value = "Lower Value Camarilla CE"
        cell_ce_l.font = header_font
        cell_ce_l.fill = header_fill
        cell_ce_l.alignment = center
        
        worksheet_lower.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)
        cell_pe_l = worksheet_lower.cell(row=1, column=4)
        cell_pe_l.value = "Lower Value Camarilla PE"
        cell_pe_l.font = header_font
        cell_pe_l.fill = header_fill
        cell_pe_l.alignment = center
        
        # Formatting
        for i, col in enumerate(df_combined_lower.columns):
            col_idx = i + 1
            series = df_combined_lower.iloc[:, i]
            max_len = max((series.apply(str).map(len).max() if not series.empty else 0), len(str(col))) + 2
            column_letter = worksheet_lower.cell(row=2, column=col_idx).column_letter
            worksheet_lower.column_dimensions[column_letter].width = max_len

        # Sheet 6: Top N Output
        metrics = [
            ('OpnIntrst', f'Top {top_n} Open Interest'),
            ('ChngInOpnIntrst', f'Top {top_n} Change in OI'),
            ('TtlTradgVol', f'Top {top_n} Volume'),
            ('TtlNbOfTxsExctd', f'Top {top_n} Transactions')
        ]
        
        sheet_name_top5 = f'Top {top_n} Output'
        workbook.create_sheet(sheet_name_top5)
        worksheet_top5 = workbook[sheet_name_top5]
        
        start_row = 1
        current_col = 0
        
        for metric, title in metrics:
            df[metric] = pd.to_numeric(df[metric], errors='coerce').fillna(0)
            
            top5_df = df.sort_values(by=metric, ascending=False).head(top_n).copy()
            
            cols_top5 = ['Symbol', 'Option_Type', 'ATM_Strike', 'Spot_Close', metric]
            top5_display = top5_df[cols_top5].copy()

            
            top5_display.to_excel(writer, sheet_name=sheet_name_top5, index=False, startrow=start_row, startcol=current_col)
            
            op_col_start = current_col + 1
            op_col_end = current_col + len(cols_top5)
            
            worksheet_top5.merge_cells(start_row=1, start_column=op_col_start, end_row=1, end_column=op_col_end)
            cell_title = worksheet_top5.cell(row=1, column=op_col_start)
            cell_title.value = title
            cell_title.font = header_font
            cell_title.fill = header_fill
            cell_title.alignment = center
            
            for i, col in enumerate(top5_display.columns):
                col_idx = op_col_start + i
                series = top5_display.iloc[:, i]
                max_len = max((series.apply(str).map(len).max() if not series.empty else 0), len(str(col))) + 2
                column_letter = worksheet_top5.cell(row=2, column=col_idx).column_letter
                worksheet_top5.column_dimensions[column_letter].width = max_len

            current_col += len(cols_top5) + 1
            
    return output.getvalue()

# Header
st.title("Camarilla Option Scanner")
st.markdown("Upload Today's and Yesterday's Bhav Copy files to generate the report.")

col1, col2 = st.columns(2)

with col1:
    st.subheader("Today's Data")
    today_file = st.file_uploader("Upload Today's Bhav Copy (ZIP)", type=['zip'], key='today')

with col2:
    st.subheader("Yesterday's Data")
    yest_file = st.file_uploader("Upload Yesterday's Bhav Copy (ZIP)", type=['zip'], key='yest')

# Option for Top N Results
st.markdown("### Report Settings")
top_n_choice = st.radio(
    "Select Number of Top Results to Display:",
    options=[5, 10],
    index=0,
    horizontal=True,
    help="Choose whether to see Top 5 or Top 10 results in the generated Excel report."
)

if st.button("SCAN & GENERATE REPORT"):
    if today_file is not None and yest_file is not None:
        try:
            with st.spinner('Processing... This may take a moment.'):
                # Save uploaded files temporarily because the scanner library expects paths
                # Or modify scanner to accept file objects. 
                # Since scanner.py uses zipfile.ZipFile(path), we can pass the file-like object directly if it supports seek!
                # Streamlit UploadedFile supports seek.
                
                # However, scanner.py prints filenames, so let's verify if it needs paths or objects.
                # scanner.py: with zipfile.ZipFile(zip_path, 'r') as z:
                # ZipFile accepts path or file-like object. So passing the UploadedFile should work!
                
                scanner = CamarillaScanner()
                df = scanner.process_data(today_file, yest_file)
                
                if df is not None and not df.empty:
                    # Generate Output Filename
                    # Try to get date from filename
                    today_filename = today_file.name
                    match = re.search(r"(\d{8})", today_filename)
                    if match:
                        date_str = match.group(1)
                    else:
                        date_str = "Report"
                    
                    output_filename = f"Camarilla Scanner {date_str}.xlsx"
                    
                    # Generate Excel
                    excel_data = generate_excel(df, today_filename, top_n_choice)
                    
                    st.success(f"Scan Complete! Found {len(df)} records.")
                    
                    # Preview Data
                    with st.expander("Preview Generated Data"):
                        st.dataframe(df.head())
                    
                    # Download Button
                    st.download_button(
                        label="Download Excel Report",
                        data=excel_data,
                        file_name=output_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                else:
                    st.error("No results found or processing failed.")
                    
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
            st.exception(e)
    else:
        st.warning("Please upload both ZIP files.")

# Instructions
with st.expander("How to use"):
    st.markdown("""
    1. Download the Bhav Copy ZIP files from NSE website for Today and Yesterday.
    2. Upload them in the respective fields above.
    3. Click 'SCAN & GENERATE REPORT'.
    4. Download the generated Excel file.
    """)
