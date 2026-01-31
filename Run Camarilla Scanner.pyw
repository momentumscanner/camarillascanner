import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import pandas as pd
from scanner import CamarillaScanner
import os
import re
import traceback

class CamarillaApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Camarilla Option Scanner")
        self.root.geometry("600x450")
        self.root.configure(bg="#f0f0f0")

        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        # Variables
        self.today_path = tk.StringVar()
        self.yest_path = tk.StringVar()
        self.status_var = tk.StringVar(value="Ready")

        self.create_widgets()

    def create_widgets(self):
        # Header
        header_frame = tk.Frame(self.root, bg="#2c3e50", height=60)
        header_frame.pack(fill=tk.X)
        header_label = tk.Label(header_frame, text="Camarilla Option Scanner", 
                                font=("Helvetica", 18, "bold"), fg="white", bg="#2c3e50")
        header_label.pack(pady=15)

        # Main Container
        main_frame = tk.Frame(self.root, bg="#f0f0f0", padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Inputs
        self.create_file_input(main_frame, "Today's Bhav Copy:", self.today_path, 0)
        self.create_file_input(main_frame, "Yesterday's Bhav Copy:", self.yest_path, 1)

        # Scan Button
        btn_frame = tk.Frame(main_frame, bg="#f0f0f0")
        btn_frame.grid(row=2, column=0, columnspan=3, pady=30)
        
        self.scan_btn = tk.Button(btn_frame, text="SCAN & GENERATE REPORT", 
                                  command=self.start_scan,
                                  font=("Helvetica", 12, "bold"), 
                                  bg="#27ae60", fg="white", 
                                  padx=20, pady=10, relief=tk.FLAT,
                                  cursor="hand2")
        self.scan_btn.pack()

        # Status Bar
        status_frame = tk.Frame(self.root, bg="#e0e0e0", height=30)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        
        self.status_label = tk.Label(status_frame, textvariable=self.status_var, 
                                     bg="#e0e0e0", fg="#333", font=("Arial", 9))
        self.status_label.pack(side=tk.LEFT, padx=10, pady=5)

    def create_file_input(self, parent, label_text, var, row):
        tk.Label(parent, text=label_text, bg="#f0f0f0", font=("Arial", 10, "bold")).grid(row=row, column=0, sticky="w", pady=10)
        
        entry = tk.Entry(parent, textvariable=var, width=40, font=("Arial", 10), state="readonly")
        entry.grid(row=row, column=1, padx=10, pady=10)
        
        btn = tk.Button(parent, text="Browse", command=lambda: self.browse_file(var), 
                        bg="#3498db", fg="white", font=("Arial", 9, "bold"), relief=tk.FLAT)
        btn.grid(row=row, column=2, padx=5, pady=10)

    def browse_file(self, var):
        filename = filedialog.askopenfilename(filetypes=[("Zip files", "*.zip")])
        if filename:
            var.set(filename)

    def start_scan(self):
        today = self.today_path.get()
        yest = self.yest_path.get()
        
        if not today or not yest:
            messagebox.showwarning("Input Error", "Please select both ZIP files.")
            return
            
        self.scan_btn.config(state=tk.DISABLED, text="Scanning...")
        self.status_var.set("Processing... Please wait.")
        
        # Run in thread to not freeze GUI
        threading.Thread(target=self.run_process, args=(today, yest)).start()

    def run_process(self, today, yest):
        try:
            scanner = CamarillaScanner()
            df = scanner.process_data(today, yest)
            
            if df is not None and not df.empty:
                # Reorder columns
                cols = list(df.columns)
                priority = ['Symbol', 'Expiry', 'ATM_Strike', 'Option_Type', 'Spot_Close']
                
                final_cols = priority + [c for c in cols if c not in priority]
                
                df_final = df[final_cols]
                
                # Filter for Sheet 2
                df_inside_full = df[df['Is_Inside_Camarilla'] == True].copy()
                
                # Split CE and PE
                # User requested ONLY: Stock Name (Symbol), Close Price (Spot_Close), Strike Price (ATM_Strike)
                cols_to_show = ['Symbol', 'Spot_Close', 'ATM_Strike']
                
                df_ce = df_inside_full[df_inside_full['Option_Type'] == 'CE'][cols_to_show].copy()
                df_pe = df_inside_full[df_inside_full['Option_Type'] == 'PE'][cols_to_show].copy()
                
                # Reset index to allow side-by-side concatenation
                df_ce.reset_index(drop=True, inplace=True)
                df_pe.reset_index(drop=True, inplace=True)
                
                # Construct combined dataframe
                df_combined = pd.concat([df_ce, df_pe], axis=1)
                
                # Dynamic Output Filename
                # Extract date from Today's filename (e.g., ...20260114...)
                basename = os.path.basename(today)
                # Look for 8 digit date pattern
                match = re.search(r"(\d{8})", basename)
                if match:
                    date_str = match.group(1)
                else:
                    date_str = "Report"
                
                output_file = f"Camarilla Scanner {date_str}.xlsx"
                
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    # Sheet 1: Main Data
                    df_final.to_excel(writer, sheet_name='Main Data', index=False)
                    
                    # Sheet 2: Inside Camarilla
                    # Write data starting from row 1 (row 2 in Excel 1-based indexing) to leave room for Top Header
                    sheet_name = 'Narrow Camarilla'
                    df_combined.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                    
                    # Get workbook and worksheet objects to format
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]
                    
                    # Add Top Level Headers (Merged)
                    from openpyxl.styles import Alignment, Font, PatternFill
                    
                    # Header Styles
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    center = Alignment(horizontal='center', vertical='center')
                    
                    # Merge and Write CE Header
                    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
                    cell_ce = worksheet.cell(row=1, column=1)
                    cell_ce.value = "Narrow Camarilla CE"
                    cell_ce.font = header_font
                    cell_ce.fill = header_fill
                    cell_ce.alignment = center
                    
                    # Merge and Write PE Header
                    worksheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)
                    cell_pe = worksheet.cell(row=1, column=4)
                    cell_pe.value = "Narrow Camarilla PE"
                    cell_pe.font = header_font
                    cell_pe.fill = header_fill
                    cell_pe.alignment = center
                    
                    # formatting columns width
                    for i, col in enumerate(df_combined.columns):
                        col_idx = i + 1 
                        # Use iloc to select by position because names might be duplicated
                        series = df_combined.iloc[:, i]
                        # Simple auto-fit estimation
                        max_len = max((series.apply(str).map(len).max() if not series.empty else 0), len(str(col))) + 2
                        column_letter = worksheet.cell(row=2, column=col_idx).column_letter
                        column_letter = worksheet.cell(row=2, column=col_idx).column_letter
                        worksheet.column_dimensions[column_letter].width = max_len

                    # Sheet 3: Inside Camarilla (H4/L4 Logic)
                    df_h4_l4_full = df[df['Is_Inside_H4_L4'] == True].copy()
                    
                    df_ce_h4 = df_h4_l4_full[df_h4_l4_full['Option_Type'] == 'CE'][cols_to_show].copy()
                    df_pe_h4 = df_h4_l4_full[df_h4_l4_full['Option_Type'] == 'PE'][cols_to_show].copy()
                    
                    df_ce_h4.reset_index(drop=True, inplace=True)
                    df_pe_h4.reset_index(drop=True, inplace=True)
                    
                    df_combined_h4 = pd.concat([df_ce_h4, df_pe_h4], axis=1)
                    
                    sheet_name_h4 = 'Inside Camarilla'
                    df_combined_h4.to_excel(writer, sheet_name=sheet_name_h4, index=False, startrow=1)
                    
                    worksheet_h4 = writer.sheets[sheet_name_h4]
                    
                    # Headers for Sheet 3
                    worksheet_h4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
                    cell_ce_h4 = worksheet_h4.cell(row=1, column=1)
                    cell_ce_h4.value = "Inside Camarilla CE"
                    cell_ce_h4.font = header_font
                    cell_ce_h4.fill = header_fill
                    cell_ce_h4.alignment = center
                    
                    worksheet_h4.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)
                    cell_pe_h4 = worksheet_h4.cell(row=1, column=4)
                    cell_pe_h4.value = "Inside Camarilla PE"
                    cell_pe_h4.font = header_font
                    cell_pe_h4.fill = header_fill
                    cell_pe_h4.alignment = center
                    
                    # Formatting columns for Sheet 3
                    for i, col in enumerate(df_combined_h4.columns):
                        col_idx = i + 1
                        series = df_combined_h4.iloc[:, i]
                        max_len = max((series.apply(str).map(len).max() if not series.empty else 0), len(str(col))) + 2
                        column_letter = worksheet_h4.cell(row=2, column=col_idx).column_letter
                        worksheet_h4.column_dimensions[column_letter].width = max_len

                    # Sheet 4: Higher Value Camarilla (Is_Higher_Value)
                    df_higher_full = df[df['Is_Higher_Value'] == True].copy()
                    
                    df_ce_higher = df_higher_full[df_higher_full['Option_Type'] == 'CE'][cols_to_show].copy()
                    df_pe_higher = df_higher_full[df_higher_full['Option_Type'] == 'PE'][cols_to_show].copy()
                    
                    df_ce_higher.reset_index(drop=True, inplace=True)
                    df_pe_higher.reset_index(drop=True, inplace=True)
                    
                    df_combined_higher = pd.concat([df_ce_higher, df_pe_higher], axis=1)
                    
                    sheet_name_higher = 'Higher Value Camarilla'
                    df_combined_higher.to_excel(writer, sheet_name=sheet_name_higher, index=False, startrow=1)
                    
                    worksheet_higher = writer.sheets[sheet_name_higher]
                    
                    # Headers
                    worksheet_higher.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
                    cell_ce_h = worksheet_higher.cell(row=1, column=1)
                    cell_ce_h.value = "Higher Value Camarilla CE"
                    cell_ce_h.font = header_font
                    cell_ce_h.fill = header_fill
                    cell_ce_h.alignment = center
                    
                    worksheet_higher.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)
                    cell_pe_h = worksheet_higher.cell(row=1, column=4)
                    cell_pe_h.value = "Higher Value Camarilla PE"
                    cell_pe_h.font = header_font
                    cell_pe_h.fill = header_fill
                    cell_pe_h.alignment = center
                    
                    # Formatting
                    for i, col in enumerate(df_combined_higher.columns):
                        col_idx = i + 1
                        series = df_combined_higher.iloc[:, i]
                        max_len = max((series.apply(str).map(len).max() if not series.empty else 0), len(str(col))) + 2
                        column_letter = worksheet_higher.cell(row=2, column=col_idx).column_letter
                        worksheet_higher.column_dimensions[column_letter].width = max_len

                    # Sheet 5: Lower Value Camarilla (Is_Lower_Value)
                    df_lower_full = df[df['Is_Lower_Value'] == True].copy()
                    
                    df_ce_lower = df_lower_full[df_lower_full['Option_Type'] == 'CE'][cols_to_show].copy()
                    df_pe_lower = df_lower_full[df_lower_full['Option_Type'] == 'PE'][cols_to_show].copy()
                    
                    df_ce_lower.reset_index(drop=True, inplace=True)
                    df_pe_lower.reset_index(drop=True, inplace=True)
                    
                    df_combined_lower = pd.concat([df_ce_lower, df_pe_lower], axis=1)
                    
                    sheet_name_lower = 'Lower Value Camarilla'
                    df_combined_lower.to_excel(writer, sheet_name=sheet_name_lower, index=False, startrow=1)
                    
                    worksheet_lower = writer.sheets[sheet_name_lower]
                    
                    # Headers
                    worksheet_lower.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
                    cell_ce_l = worksheet_lower.cell(row=1, column=1)
                    cell_ce_l.value = "Lower Value Camarilla CE"
                    cell_ce_l.font = header_font
                    cell_ce_l.fill = header_fill
                    cell_ce_l.alignment = center
                    
                    worksheet_lower.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)
                    cell_pe_l = worksheet_lower.cell(row=1, column=4)
                    cell_pe_l.value = "Lower Value Camarilla PE"
                    cell_pe_l.font = header_font
                    cell_pe_l.fill = header_fill
                    cell_pe_l.alignment = center
                    
                    # Formatting
                    for i, col in enumerate(df_combined_lower.columns):
                        col_idx = i + 1
                        series = df_combined_lower.iloc[:, i]
                        max_len = max((series.apply(str).map(len).max() if not series.empty else 0), len(str(col))) + 2
                        column_letter = worksheet_lower.cell(row=2, column=col_idx).column_letter

                    # Sheet 6: Top 5 Output
                    # Metrics to process
                    metrics = [
                        ('OpnIntrst', 'Top 5 Open Interest'),
                        ('ChngInOpnIntrst', 'Top 5 Change in OI'),
                        ('TtlTradgVol', 'Top 5 Volume'),
                        ('TtlNbOfTxsExctd', 'Top 5 Transactions')
                    ]
                    
                    sheet_name_top5 = 'Top 5 Output'
                    # Initialize sheet by writing a placeholder or just accessing it via openpyxl if we want custom positions
                    # But pandas to_excel needs to write something to create the sheet or use the writer
                    # We will write each table at specific offsets
                    
                    start_row = 1
                    start_col = 0 # 0-based for pandas to_excel? No, pandas uses startrow/startcol.
                    
                    # Create sheet first
                    workbook.create_sheet(sheet_name_top5)
                    worksheet_top5 = workbook[sheet_name_top5]
                    
                    current_col = 0
                    
                    for metric, title in metrics:
                        # Sort and take top 5
                        # Ensure numeric
                        df[metric] = pd.to_numeric(df[metric], errors='coerce').fillna(0)
                        
                        top5_df = df.sort_values(by=metric, ascending=False).head(5).copy()
                        
                        # Select columns
                        cols_top5 = ['Symbol', 'Option_Type', 'ATM_Strike', 'Spot_Close', metric]
                        top5_display = top5_df[cols_top5].copy()
                        
                        # Write to Excel
                        # pandas to_excel startcol is 0-based.
                        top5_display.to_excel(writer, sheet_name=sheet_name_top5, index=False, startrow=start_row, startcol=current_col)
                        
                        # Add Header
                        # Create a merged header above the table
                        # Pandas writes header at start_row. So we need the title at start_row - 1.
                        # Wait, startrow=1 means row 2. So row 1 is free.
                        
                        # Merge cells for title
                        # openpyxl uses 1-based indexing for rows/cols
                        op_col_start = current_col + 1
                        op_col_end = current_col + len(cols_top5)
                        
                        worksheet_top5.merge_cells(start_row=1, start_column=op_col_start, end_row=1, end_column=op_col_end)
                        cell_title = worksheet_top5.cell(row=1, column=op_col_start)
                        cell_title.value = title
                        cell_title.font = header_font
                        cell_title.fill = header_fill
                        cell_title.alignment = center
                        
                        # Formatting columns
                        for i, col in enumerate(top5_display.columns):
                            col_idx = op_col_start + i
                            series = top5_display.iloc[:, i]
                            max_len = max((series.apply(str).map(len).max() if not series.empty else 0), len(str(col))) + 2
                            column_letter = worksheet_top5.cell(row=2, column=col_idx).column_letter
                            worksheet_top5.column_dimensions[column_letter].width = max_len

                        # Move to next block (width + 1 column gap)
                        current_col += len(cols_top5) + 1

                self.root.after(0, lambda: self.scan_success(output_file))
            else:
                self.root.after(0, lambda: self.scan_fail("No results found."))
        
        except PermissionError:
            self.root.after(0, lambda: self.scan_fail(f"Permission Denied!\nPlease close '{output_file}' and try again."))
                
        except Exception as e:
            err_msg = str(e)
            traceback_str = traceback.format_exc()
            print(f"Error: {err_msg}") # Print to console as well
            print(traceback_str)
            self.root.after(0, lambda: self.scan_fail(f"{err_msg}\n\n{traceback_str}"))

    def scan_success(self, filename):
        self.status_var.set(f"Completed! Saved to {filename}")
        self.scan_btn.config(state=tk.NORMAL, text="SCAN & GENERATE REPORT")
        messagebox.showinfo("Success", f"Scan Complete!\nFile saved as: {filename}")

    def scan_fail(self, error):
        self.status_var.set("Failed.")
        self.scan_btn.config(state=tk.NORMAL, text="SCAN & GENERATE REPORT")
        # messagebox.showerror("Error", f"Processing Failed:\n{error}") # Avoid popup here if relying on on-screen status? No, popup is better for errors.
        if "Permission Denied" in str(error):
             messagebox.showwarning("File Open", error)
        else:
             messagebox.showerror("Error", f"Processing Failed:\n{error}")

if __name__ == "__main__":
    root = tk.Tk()
    app = CamarillaApp(root)
    root.mainloop()
