
import pandas as pd
import os

def verify_top5_generation():
    # Mock Data
    data = {
        'Symbol': ['A', 'B', 'C', 'D', 'E', 'F', 'G'],
        'Option_Type': ['CE', 'PE', 'CE', 'PE', 'CE', 'PE', 'CE'],
        'ATM_Strike': [100, 200, 300, 400, 500, 600, 700],
        'Spot_Close': [101, 201, 301, 401, 501, 601, 701],
        'OpnIntrst': [10, 50, 20, 60, 30, 70, 40], # Top 5 should be F(70), D(60), B(50), G(40), E(30)
        'ChngInOpnIntrst': [1, 2, 3, 4, 5, 6, 7],
        'TtlTradgVol': [1000, 2000, 3000, 4000, 5000, 6000, 7000],
        'TtlNbOfTxsExctd': [5, 5, 5, 5, 5, 5, 5]
    }
    
    df = pd.DataFrame(data)
    
    output_file = "Verify_Top5.xlsx"
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Main Data', index=False)
        
        # --- LOGIC START ---
        metrics = [
            ('OpnIntrst', 'Top 5 Open Interest'),
            ('ChngInOpnIntrst', 'Top 5 Change in OI'),
            ('TtlTradgVol', 'Top 5 Volume'),
            ('TtlNbOfTxsExctd', 'Top 5 Transactions')
        ]
        
        sheet_name_top5 = 'Top 5 Output'
        workbook = writer.book
        workbook.create_sheet(sheet_name_top5)
        worksheet_top5 = workbook[sheet_name_top5]
        
        from openpyxl.styles import Alignment, Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center = Alignment(horizontal='center', vertical='center')

        start_row = 1
        current_col = 0
        
        for metric, title in metrics:
            print(f"Processing {metric}...")
            # Sort and take top 5
            df[metric] = pd.to_numeric(df[metric], errors='coerce').fillna(0)
            
            top5_df = df.sort_values(by=metric, ascending=False).head(5).copy()
            
            # Select columns
            cols_top5 = ['Symbol', 'Option_Type', 'ATM_Strike', 'Spot_Close', metric]
            top5_display = top5_df[cols_top5].copy()
            
            # Write to Excel
            top5_display.to_excel(writer, sheet_name=sheet_name_top5, index=False, startrow=start_row, startcol=current_col)
            
            # Add Header
            op_col_start = current_col + 1
            op_col_end = current_col + len(cols_top5)
            
            worksheet_top5.merge_cells(start_row=1, start_column=op_col_start, end_row=1, end_column=op_col_end)
            cell_title = worksheet_top5.cell(row=1, column=op_col_start)
            cell_title.value = title
            cell_title.font = header_font
            cell_title.fill = header_fill
            cell_title.alignment = center
            
            # Formatting columns
            for i, col in enumerate(top5_display.columns):
                col_idx = op_col_start + i
                series = top5_display.iloc[:, i]
                max_len = max((series.apply(str).map(len).max() if not series.empty else 0), len(str(col))) + 2
                column_letter = worksheet_top5.cell(row=2, column=col_idx).column_letter
                worksheet_top5.column_dimensions[column_letter].width = max_len

            # Move to next block (width + 1 column gap)
            current_col += len(cols_top5) + 1
            
        # --- LOGIC END ---
        
    print(f"Verification successful! Saved {output_file}")
    
if __name__ == "__main__":
    try:
        verify_top5_generation()
    except Exception as e:
        print(f"Output Generation Failed: {e}")
